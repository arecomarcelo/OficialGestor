import tkinter as tk
from tkinter import messagebox
from PIL import Image, ImageTk
from vars import QuebraLinha, nome_aplicacao, celular_suporte, email_suporte, caminhoArquivos, dtAtualFormatadaArquivo
from imagens import icone_oficial
from enumeradores import Direcao
from openpyxl import Workbook
from funcoes import ApagarArquivo

def MontaTela(cor_fundo, titulo, principal = False):
    if principal:
        tela = tk.Tk()
    else:
        tela = tk.Toplevel()    
        
    tela.geometry('1960x800+0+0')
    tela.state('zoomed')
    tela.title(f"{nome_aplicacao} - {titulo}")
    tela['bg'] = cor_fundo

    icone_image = Image.open(icone_oficial)
    icone_photo = ImageTk.PhotoImage(icone_image)
    tela.iconphoto(False, icone_photo)

    # image = Image.open(fundo)
    # tkimage = ImageTk.PhotoImage(image)
    # label = tk.Label(tela, image=tkimage, bd=0, highlightthickness=0)
    # label.pack(padx=0, pady=0)
    # tk.Label(tela, image=tkimage).pack() 
        
    return tela

def PosicionaBotao (tela, botao_anterior, botao_atual, direcao = Direcao.DIREITA):
    tela.update_idletasks()
    
    width = botao_anterior.winfo_width() + 5
    
    if direcao.name == "ESQUERDA":
        posx = botao_anterior.winfo_x() + width
    else:
        posx = botao_anterior.winfo_x() - width
    
    posy = botao_anterior.winfo_y()    
    
    botao_atual.place(x = posx, y = posy, width = 100)   
    


    
def boletos_planilha():
    exec(open("G:\\Meu Drive\\TI Oficial\\Projetos\\Cobranca\\Envio_Whatsapp\\BOT\\main.py", encoding="utf-8").read(),locals())

def relatorio_conferencia():
    exec(open("main\\relatorio_conferencia.py", encoding="utf-8").read(),locals())
    
def estoque_listagem():
    exec(open("main\estoque_listagem.py", encoding="utf-8").read(),locals())
    
def suporte():
    messagebox.showinfo("Contatos de Suporte", f"{QuebraLinha}Whatsapp: {celular_suporte}{QuebraLinha}E-Mail: {email_suporte}")
    
def sobre():
    messagebox.showinfo("Sobre", nome_aplicacao)
    
def sair(tela):
    nome_tela = tela.title()
    
    var_sair = messagebox.askyesno("Sair", f"Tem certeza que deseja sair de {nome_tela} ?")
    if var_sair:
        tela.destroy()  
        
def fechar_tela(tela, event=None):
    tela.destroy()
    
def ExportaDadosGrid(tela, treeview):
    nomearquivo = f"{caminhoArquivos}Relatório {dtAtualFormatadaArquivo}.xlsx"
    
    ApagarArquivo(nomearquivo)
    
    # Cria um novo Workbook e seleciona a planilha ativa
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados do Treeview"
    
    # Obtém os cabeçalhos das colunas do Treeview
    colunas = treeview["columns"]
    
    # Escreve os cabeçalhos na primeira linha da planilha
    for col_index, col_name in enumerate(colunas, start=1):
        ws.cell(row=1, column=col_index, value=treeview.heading(col_name)["text"])
    
    # Itera sobre os itens do Treeview e escreve os valores na planilha
    for row_index, item_id in enumerate(treeview.get_children(), start=2):
        valores = treeview.item(item_id, "values")
        for col_index, valor in enumerate(valores, start=1):
            ws.cell(row=row_index, column=col_index, value=valor)
    
    # Salva o arquivo Excel
    wb.save(nomearquivo)
    tk.messagebox.showinfo (nome_aplicacao, "Relatório Gerado com Sucesso",parent = tela)